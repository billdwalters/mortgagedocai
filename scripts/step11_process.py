#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import subprocess
import uuid
from pathlib import Path
from typing import Any, Dict, List, Tuple

import torch
from sentence_transformers import SentenceTransformer
from qdrant_client import QdrantClient
from qdrant_client.http.models import Distance, VectorParams, PointStruct

from lib import (
    DEFAULT_TENANT,
    ContractError,
    NAS_INGEST,
    EMBED_DIM,
    EMBED_MODEL_NAME,
    preflight_mount_contract,
    build_run_context,
    qdrant_collection_name,
    normalize_chunk_text,
    chunk_id,
    chunk_text_hash,
    atomic_write_json,
    atomic_write_text,
    atomic_rename_dir,
    ensure_dir,
    sha256_file,
    utc_run_id,
)

# ---------------------------------------------------------------------
# Deterministic Qdrant point ID (UUIDv5 from chunk_id)
# ---------------------------------------------------------------------
_UUID_NAMESPACE = uuid.UUID("b1f0b5d8-2a54-4d7f-8f1e-9f9b0b3a9b1a")


def point_id_from_chunk_id(chunk_id_hex: str) -> str:
    return str(uuid.uuid5(_UUID_NAMESPACE, chunk_id_hex))


# ---------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------
def parse_args(argv=None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Step 11 — Process (v1)")
    p.add_argument("--tenant-id", default=DEFAULT_TENANT)
    p.add_argument("--loan-id", required=True)
    p.add_argument("--run-id", required=True)

    p.add_argument("--qdrant-url", default="http://localhost:6333")
    p.add_argument("--embedding-model", default=EMBED_MODEL_NAME)
    p.add_argument("--embedding-dim", type=int, default=EMBED_DIM)
    p.add_argument("--embedding-device", choices=["cpu", "cuda"], default=None)
    p.add_argument("--batch-size", type=int, default=32)

    p.add_argument("--chunk-target-chars", type=int, default=4500)
    p.add_argument("--chunk-max-chars", type=int, default=6000)
    p.add_argument("--chunk-overlap-chars", type=int, default=800)
    p.add_argument("--min-chunk-chars", type=int, default=900)

    p.add_argument("--dense-chunk-target-chars", type=int, default=2400)
    p.add_argument("--dense-chunk-max-chars", type=int, default=3400)
    p.add_argument("--dense-chunk-overlap-chars", type=int, default=350)

    p.add_argument("--ocr-threshold-chars", type=int, default=400)
    return p.parse_args(argv)


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------
def _extract_pdf_pages_text(pdf_path: Path) -> List[str]:
    from pypdf import PdfReader

    reader = PdfReader(str(pdf_path))
    if reader.is_encrypted:
        raise RuntimeError("encrypted_pdf")

    pages: List[str] = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return pages


def _extract_docx_text(path: Path) -> str:
    """Extract text from .docx; deterministic (paragraphs joined with newlines)."""
    from docx import Document

    doc = Document(str(path))
    parts: List[str] = []
    for para in doc.paragraphs:
        t = (para.text or "").strip()
        if t:
            parts.append(t)
    return "\n".join(parts)


def _extract_xlsx_text(path: Path) -> str:
    """Extract text from .xlsx; deterministic (sheets by name, rows tab-separated)."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: List[str] = []
    for sheet_name in sorted(wb.sheetnames):
        parts.append(f"Sheet: {sheet_name}")
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in (row or [])]
            parts.append("\t".join(cells))
        parts.append("")
    wb.close()
    return "\n".join(parts).strip()


def _looks_dense_doc(filename: str) -> bool:
    name = filename.lower()
    return any(
        k in name
        for k in ["w2", "1099", "paystub", "statement", "1003", "uw", "credit", "tax", "id"]
    )


def _chunk_page_text(
    text: str, target: int, maxc: int, overlap: int, minc: int
) -> List[str]:
    t = normalize_chunk_text(text)
    if not t:
        return []

    chunks: List[str] = []
    i = 0
    n = len(t)

    while i < n:
        end = min(i + target, n)
        end = min(end, i + maxc)
        chunks.append(t[i:end])
        if end >= n:
            break
        i = max(0, end - overlap)

    merged: List[str] = []
    buf = ""
    for c in chunks:
        if not buf:
            buf = c
        elif len(buf) < minc:
            buf = (buf + "\n" + c).strip()
        else:
            merged.append(buf)
            buf = c
    if buf:
        merged.append(buf)

    return merged


def _ensure_collection(client: QdrantClient, collection: str) -> None:
    try:
        info = client.get_collection(collection)
        vp = info.config.params.vectors
        if isinstance(vp, dict):
            vp = list(vp.values())[0]
        if int(vp.size) != EMBED_DIM:
            raise ContractError("Qdrant collection dim mismatch")
        if vp.distance != Distance.COSINE:
            raise ContractError("Qdrant collection distance mismatch")
    except Exception:
        client.recreate_collection(
            collection_name=collection,
            vectors_config=VectorParams(size=EMBED_DIM, distance=Distance.COSINE),
        )


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------
def main(argv=None) -> None:
    args = parse_args(argv)
    preflight_mount_contract()

    if args.embedding_dim != EMBED_DIM:
        raise ContractError("v1 requires embedding_dim=1024")

    ctx = build_run_context(args.tenant_id, args.loan_id, args.run_id)

    ingest_root = NAS_INGEST / "tenants" / args.tenant_id / "loans" / args.loan_id
    manifest_path = ingest_root / "_meta" / "intake_manifest.json"
    if not manifest_path.exists():
        raise ContractError(f"Missing intake_manifest.json: {manifest_path}")

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    files = manifest.get("files", [])

    run_staging = ctx.chunk_staging_run_root
    run_final = ctx.chunk_final_run_root
    ensure_dir(run_staging)

    for d in ["_meta", "text", "ocr", "chunks", "embeddings/qdrant"]:
        ensure_dir(run_staging / d)

    device = args.embedding_device or ("cuda" if torch.cuda.is_available() else "cpu")
    model = SentenceTransformer(args.embedding_model, device=device)

    qdrant = QdrantClient(url=args.qdrant_url)
    collection = qdrant_collection_name(args.tenant_id)
    _ensure_collection(qdrant, collection)
    atomic_write_text(run_staging / "embeddings/qdrant/collection_name.txt", collection)

    batch_texts: List[str] = []
    batch_meta: List[Tuple[str, Dict[str, Any]]] = []
    upserts = 0
    total_chunks = 0
    skipped_encrypted_count = 0

    # Accumulate chunks per document for artifacts
    # Key: document_id -> list of chunk record dicts
    doc_chunks: Dict[str, List[Dict[str, Any]]] = {}

    def flush_batch() -> None:
        nonlocal upserts, batch_texts, batch_meta
        if not batch_meta:
            return

        vecs = model.encode(
            batch_texts, normalize_embeddings=True, convert_to_numpy=True
        )
        points = [
            PointStruct(id=pid, vector=vecs[i].tolist(), payload=payload)
            for i, (pid, payload) in enumerate(batch_meta)
        ]
        qdrant.upsert(collection_name=collection, points=points, wait=True)
        upserts += len(points)
        batch_texts.clear()
        batch_meta.clear()

    for f in files:
        stored_rel = f["stored_relative_path"]
        document_id = f["document_id"]
        staged_path = ingest_root / stored_rel

        if sha256_file(staged_path) != f["sha256"]:
            raise ContractError(f"Hash mismatch: {staged_path}")

        suffix = staged_path.suffix.lower()
        pages: List[str] = []

        if suffix == ".pdf":
            try:
                pages = _extract_pdf_pages_text(staged_path)
            except Exception as e:
                print(f"[WARN] Skipping unreadable/encrypted PDF: {staged_path} ({e})")
                skipped_encrypted_count += 1
                continue
        elif suffix == ".docx":
            try:
                text = _extract_docx_text(staged_path)
                text_path = run_staging / "text" / f"{document_id}.txt"
                atomic_write_text(text_path, text)
                pages = [text] if text.strip() else []
            except Exception as e:
                print(f"[WARN] Skipping unreadable DOCX: {staged_path} ({e})")
                continue
        elif suffix == ".xlsx":
            try:
                text = _extract_xlsx_text(staged_path)
                text_path = run_staging / "text" / f"{document_id}.txt"
                atomic_write_text(text_path, text)
                pages = [text] if text.strip() else []
            except Exception as e:
                print(f"[WARN] Skipping unreadable XLSX: {staged_path} ({e})")
                continue
        else:
            continue

        if not pages:
            continue

        dense = _looks_dense_doc(staged_path.name)

        if dense:
            target, maxc, overlap, minc = (
                args.dense_chunk_target_chars,
                args.dense_chunk_max_chars,
                args.dense_chunk_overlap_chars,
                max(1, args.min_chunk_chars // 2),
            )
        else:
            target, maxc, overlap, minc = (
                args.chunk_target_chars,
                args.chunk_max_chars,
                args.chunk_overlap_chars,
                args.min_chunk_chars,
            )

        for page_i, page_text in enumerate(pages, start=1):
            for cidx, chunk in enumerate(
                _chunk_page_text(page_text, target, maxc, overlap, minc)
            ):
                cid = chunk_id(document_id, page_i, page_i, cidx, chunk)
                pid = point_id_from_chunk_id(cid)
                payload = {
                    "tenant_id": args.tenant_id,
                    "loan_id": args.loan_id,
                    "run_id": args.run_id,
                    "document_id": document_id,
                    "file_relpath": stored_rel,
                    "page_start": page_i,
                    "page_end": page_i,
                    "chunk_index": cidx,
                    "chunk_id": cid,
                }
                batch_texts.append("passage: " + chunk)
                batch_meta.append((pid, payload))
                total_chunks += 1

                # Accumulate chunk record for artifacts (text WITHOUT "passage: " prefix)
                chunk_record = {
                    "chunk_id": cid,
                    "document_id": document_id,
                    "file_relpath": stored_rel,
                    "page_start": page_i,
                    "page_end": page_i,
                    "chunk_index": cidx,
                    "text": chunk,
                    "text_norm_sha256": chunk_text_hash(chunk),
                }
                if document_id not in doc_chunks:
                    doc_chunks[document_id] = []
                doc_chunks[document_id].append(chunk_record)

                if len(batch_meta) >= args.batch_size:
                    flush_batch()

    flush_batch()

    # ---------------------------------------------------------------
    # PART 1A: Write chunks.jsonl per document
    # ---------------------------------------------------------------
    for document_id, chunk_records in doc_chunks.items():
        doc_chunk_dir = run_staging / "chunks" / document_id
        ensure_dir(doc_chunk_dir)
        lines = [json.dumps(rec, ensure_ascii=False) for rec in chunk_records]
        atomic_write_text(doc_chunk_dir / "chunks.jsonl", "\n".join(lines) + "\n")

    # ---------------------------------------------------------------
    # PART 1B: Write chunk_map.json per document
    # ---------------------------------------------------------------
    for document_id, chunk_records in doc_chunks.items():
        doc_chunk_dir = run_staging / "chunks" / document_id
        chunk_map: Dict[str, Dict[str, Any]] = {}
        for rec in chunk_records:
            chunk_map[rec["chunk_id"]] = {
                "document_id": rec["document_id"],
                "file_relpath": rec["file_relpath"],
                "page_start": rec["page_start"],
                "page_end": rec["page_end"],
                "chunk_index": rec["chunk_index"],
            }
        atomic_write_json(doc_chunk_dir / "chunk_map.json", chunk_map)

    # ---------------------------------------------------------------
    # PART 1C: Write processing_run.json
    # ---------------------------------------------------------------
    processing_meta = {
        "tenant_id": args.tenant_id,
        "loan_id": args.loan_id,
        "run_id": args.run_id,
        "embedding_model": args.embedding_model,
        "embedding_dim": args.embedding_dim,
        "distance_metric": "cosine",
        "normalize": True,
        "device": device,
        "chunker": {
            "chunk_target_chars": args.chunk_target_chars,
            "chunk_max_chars": args.chunk_max_chars,
            "chunk_overlap_chars": args.chunk_overlap_chars,
            "min_chunk_chars": args.min_chunk_chars,
            "dense_chunk_target_chars": args.dense_chunk_target_chars,
            "dense_chunk_max_chars": args.dense_chunk_max_chars,
            "dense_chunk_overlap_chars": args.dense_chunk_overlap_chars,
            "ocr_threshold_chars": args.ocr_threshold_chars,
        },
        "total_chunks": total_chunks,
        "upserts": upserts,
        "skipped_encrypted_count": skipped_encrypted_count,
        "documents_processed": len(doc_chunks),
        "qdrant_collection": collection,
    }
    atomic_write_json(run_staging / "_meta" / "processing_run.json", processing_meta)

    # ---------------------------------------------------------------
    # PART 3: Contract assertions before atomic publish
    # ---------------------------------------------------------------
    if not (run_staging / "_meta" / "processing_run.json").exists():
        raise ContractError("Step11 missing required artifacts: processing_run.json; aborting publish")

    chunks_jsonl_files = list((run_staging / "chunks").rglob("chunks.jsonl"))
    if not chunks_jsonl_files:
        raise ContractError("Step11 missing required artifacts: no chunks.jsonl found; aborting publish")

    # Rerun policy: if run_final already exists (prior attempt), remove before atomic publish
    overwrite = False
    if run_final.exists():
        import shutil
        shutil.rmtree(run_final)
        overwrite = True

    ensure_dir(run_final.parent)
    atomic_rename_dir(run_staging, run_final)
    overwrite_msg = " (overwrote prior run)" if overwrite else ""
    print(f"✓ Step11 complete: {run_final} (upserts={upserts}, chunks={total_chunks}, docs={len(doc_chunks)}, skipped_encrypted={skipped_encrypted_count}){overwrite_msg}")


if __name__ == "__main__":
    main()
