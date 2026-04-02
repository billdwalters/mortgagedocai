"""Tests for formfill.py — form registry, field mappings, and filler logic."""
import json
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent))

from formfill import (
    FORM_TEMPLATES,
    CellType,
    FieldMapping,
    FormTemplate,
    _load_source_data,
    _resolve_json_path,
    fill_form,
)


# ---------------------------------------------------------------------------
# Registry tests
# ---------------------------------------------------------------------------

def test_registry_has_eight_templates():
    assert len(FORM_TEMPLATES) == 8


def test_registry_keys_match_template_ids():
    for key, tmpl in FORM_TEMPLATES.items():
        assert key == tmpl.template_id


def test_registry_income_calc_w2_exists():
    t = FORM_TEMPLATES["income_calc_w2"]
    assert t.display_name == "Income Calc (W2)"
    assert t.category == "Income"
    assert t.filename == "income_calc_w2.xlsx"
    assert len(t.mappings) >= 9


def test_registry_fha_max_mortgage_exists():
    t = FORM_TEMPLATES["fha_max_mortgage_calc"]
    assert t.display_name == "FHA Max Mortgage Calc"
    assert t.category == "FHA"


def test_registry_va_irrrl_exists():
    t = FORM_TEMPLATES["va_irrrl_recoupment_calc"]
    assert t.display_name == "VA IRRRL Recoupment Calc"
    assert t.category == "VA"


def test_registry_fha_max_mtg_initial_exists():
    t = FORM_TEMPLATES["fha_max_mtg_initial"]
    assert t.display_name == "FHA Max Mtg Worksheet (Initial)"
    assert t.category == "FHA"
    assert len(t.mappings) >= 1


def test_registry_va_irrrl_worksheet_exists():
    t = FORM_TEMPLATES["va_irrrl_worksheet"]
    assert t.display_name == "VA-IRRRL Worksheet"
    assert t.category == "VA"
    assert len(t.mappings) >= 2


def test_registry_va_max_mortgage_exists():
    t = FORM_TEMPLATES["va_max_mortgage"]
    assert t.display_name == "VA Max Mortgage Worksheet"
    assert t.category == "VA"
    assert len(t.mappings) >= 1


def test_registry_income_calc_uwm_exists():
    t = FORM_TEMPLATES["income_calc_uwm"]
    assert t.display_name == "Income Calc (UWM)"
    assert t.category == "Income"
    assert len(t.mappings) >= 10


def test_registry_bank_stmt_loan_calc_exists():
    t = FORM_TEMPLATES["bank_stmt_loan_calc"]
    assert t.display_name == "Bank Statement Loan Calculator"
    assert t.category == "Income"
    assert len(t.mappings) >= 3


def test_all_templates_have_xlsx_files():
    forms_dir = Path(__file__).resolve().parent.parent / "webui" / "forms"
    for tmpl in FORM_TEMPLATES.values():
        xlsx = forms_dir / tmpl.filename
        assert xlsx.is_file(), f"Missing template file: {xlsx}"


# ---------------------------------------------------------------------------
# _resolve_json_path tests
# ---------------------------------------------------------------------------

def test_resolve_json_path_simple():
    data = {"monthly_income_total": {"value": 6500.0}}
    assert _resolve_json_path(data, "monthly_income_total.value") == 6500.0


def test_resolve_json_path_nested():
    data = {"decision_primary": {"status": "PASS"}}
    assert _resolve_json_path(data, "decision_primary.status") == "PASS"


def test_resolve_json_path_missing_returns_none():
    data = {"a": {"b": 1}}
    assert _resolve_json_path(data, "a.c.d") is None


def test_resolve_json_path_list_index():
    data = {"items": [{"name": "first"}, {"name": "second"}]}
    assert _resolve_json_path(data, "items.0.name") == "first"


def test_resolve_json_path_top_level():
    data = {"confidence": 0.85}
    assert _resolve_json_path(data, "confidence") == 0.85


# ---------------------------------------------------------------------------
# _load_source_data tests
# ---------------------------------------------------------------------------

def test_load_source_data_all_present(tmp_path):
    profiles = tmp_path / "profiles"
    (profiles / "income_analysis").mkdir(parents=True)
    (profiles / "uw_decision").mkdir(parents=True)
    (profiles / "uw_conditions").mkdir(parents=True)
    (profiles / "income_analysis" / "income_analysis.json").write_text(
        json.dumps({"monthly_income_total": {"value": 5000}})
    )
    (profiles / "income_analysis" / "dti.json").write_text(
        json.dumps({"front_end_dti": 0.28})
    )
    (profiles / "uw_decision" / "decision.json").write_text(
        json.dumps({"decision_primary": {"status": "PASS"}})
    )
    (profiles / "uw_conditions" / "conditions.json").write_text(
        json.dumps({"conditions": []})
    )
    data = _load_source_data(profiles)
    assert data["income_analysis"]["monthly_income_total"]["value"] == 5000
    assert data["dti"]["front_end_dti"] == 0.28
    assert data["decision"]["decision_primary"]["status"] == "PASS"
    assert data["conditions"]["conditions"] == []


def test_load_source_data_missing_files(tmp_path):
    profiles = tmp_path / "profiles"
    profiles.mkdir(parents=True)
    data = _load_source_data(profiles)
    assert data["income_analysis"] == {}
    assert data["dti"] == {}
    assert data["decision"] == {}
    assert data["conditions"] == {}


# ---------------------------------------------------------------------------
# fill_form tests
# ---------------------------------------------------------------------------

def _make_profiles_dir(tmp_path):
    """Create a profiles dir with sample pipeline output data."""
    profiles = tmp_path / "profiles"
    (profiles / "income_analysis").mkdir(parents=True)
    (profiles / "uw_decision").mkdir(parents=True)
    (profiles / "uw_conditions").mkdir(parents=True)
    (profiles / "income_analysis" / "income_analysis.json").write_text(json.dumps({
        "borrowers": [
            {"name": "John Smith", "role": "primary", "employer": "Acme Corp",
             "employment_type": "W2", "citations": []},
        ],
        "monthly_income_total": {
            "value": 6500.0,
            "components": [
                {"period_total": 58500.0, "period_months": 9.0, "monthly_equivalent": 6500.0},
                {"period_total": 30000.0, "period_months": 12.0, "monthly_equivalent": 2500.0},
            ],
        },
        "monthly_liabilities_total": {"value": 1250.0},
        "proposed_pitia": {"value": 2100.0},
        "income_items": [
            {"description": "W-2 Income", "amount": 6500.0, "frequency": "monthly"}
        ],
    }))
    (profiles / "income_analysis" / "dti.json").write_text(json.dumps({
        "front_end_dti": 0.3231,
        "back_end_dti": 0.5154,
        "housing_payment_used": 2100.0,
        "monthly_income_total": 6500.0,
        "monthly_debt_total": 1250.0,
    }))
    (profiles / "uw_decision" / "decision.json").write_text(json.dumps({
        "decision_primary": {"status": "FAIL", "reasons": []},
    }))
    (profiles / "uw_conditions" / "conditions.json").write_text(json.dumps({
        "conditions": [{"description": "Provide bank statements"}],
    }))
    return profiles


def test_fill_form_returns_audit_dict(tmp_path):
    profiles = _make_profiles_dir(tmp_path)
    out = tmp_path / "output" / "income_calc_w2.xlsx"
    result = fill_form("income_calc_w2", profiles, out, loan_id="12345", run_id="2026-01-01T120000Z")
    assert isinstance(result, dict)
    assert result["template_id"] == "income_calc_w2"
    assert "cells_filled" in result
    assert "cells_total" in result
    assert "skipped_fields" in result
    assert out.is_file()


def test_fill_form_creates_output_dir(tmp_path):
    profiles = _make_profiles_dir(tmp_path)
    out = tmp_path / "deep" / "nested" / "output.xlsx"
    fill_form("income_calc_w2", profiles, out, loan_id="12345", run_id="2026-01-01T120000Z")
    assert out.is_file()


def test_fill_form_preserves_formulas(tmp_path):
    import openpyxl
    profiles = _make_profiles_dir(tmp_path)
    out = tmp_path / "output" / "income_calc_w2.xlsx"
    fill_form("income_calc_w2", profiles, out, loan_id="12345", run_id="2026-01-01T120000Z")
    wb = openpyxl.load_workbook(str(out), data_only=False)
    ws = wb[wb.sheetnames[0]]
    # J10 should still have the formula =C10*G10*52/12
    assert ws["J10"].value is not None
    val = str(ws["J10"].value)
    assert val.startswith("="), f"Expected formula in J10, got {val}"


def test_fill_form_invalid_template_raises(tmp_path):
    profiles = _make_profiles_dir(tmp_path)
    out = tmp_path / "output" / "bad.xlsx"
    with pytest.raises(ValueError, match="Unknown template_id"):
        fill_form("nonexistent_template", profiles, out, loan_id="1", run_id="x")


def test_fill_form_skips_missing_values(tmp_path):
    """When pipeline data is missing, cells are left empty and logged in skipped_fields."""
    profiles = tmp_path / "profiles"
    profiles.mkdir(parents=True)
    # Empty profiles — no data at all
    out = tmp_path / "output" / "income_calc_w2.xlsx"
    result = fill_form("income_calc_w2", profiles, out, loan_id="1", run_id="x")
    assert result["cells_filled"] == 0
    assert len(result["skipped_fields"]) == result["cells_total"]


def test_fill_form_w2_fills_borrower_name(tmp_path):
    """Income Calc W2 should fill borrower name from income_analysis.borrowers."""
    import openpyxl
    profiles = _make_profiles_dir(tmp_path)
    out = tmp_path / "output" / "income_calc_w2.xlsx"
    result = fill_form("income_calc_w2", profiles, out, loan_id="12345", run_id="2026-01-01T120000Z")
    assert result["cells_filled"] >= 5  # borrower, employer, income components, salary, PITIA, debt
    wb = openpyxl.load_workbook(str(out))
    ws = wb[wb.sheetnames[0]]
    assert ws["C4"].value == "John Smith"
    assert ws["C5"].value == "Acme Corp"


def test_fill_form_uwm_fills_multiple_sheets(tmp_path):
    """Income Calc UWM should fill borrower name on multiple sheets."""
    import openpyxl
    profiles = _make_profiles_dir(tmp_path)
    out = tmp_path / "output" / "income_calc_uwm.xlsm"
    result = fill_form("income_calc_uwm", profiles, out, loan_id="12345", run_id="2026-01-01T120000Z")
    assert result["cells_filled"] >= 5  # borrower + employer on multiple sheets + monthly rate
    wb = openpyxl.load_workbook(str(out), keep_vba=True)
    assert wb["Full Time Monthly"]["C7"].value == "John Smith"
    assert wb["Full Time Monthly"]["C8"].value == "Acme Corp"
    assert wb["Full Time Monthly"]["E11"].value == 6500.0
    assert wb["Income Calculator"]["C4"].value == "John Smith"


def test_fill_form_bank_stmt_fills_pitia(tmp_path):
    """Bank Statement Loan Calc should fill Subject PITIA on Reserve Calculator."""
    import openpyxl
    profiles = _make_profiles_dir(tmp_path)
    out = tmp_path / "output" / "bank_stmt.xlsm"
    result = fill_form("bank_stmt_loan_calc", profiles, out, loan_id="12345", run_id="2026-01-01T120000Z")
    assert result["cells_filled"] >= 2  # business name + PITIA at minimum
    wb = openpyxl.load_workbook(str(out), keep_vba=True)
    assert wb["Reserve Calculator"]["C7"].value == 2100.0


def test_fill_form_fha_max_mortgage_fills_three_sheets(tmp_path):
    """FHA Max Mortgage Calc should fill PITIA on Simple, Streamline, and R&T sheets."""
    import openpyxl
    profiles = _make_profiles_dir(tmp_path)
    out = tmp_path / "output" / "fha_max.xlsx"
    result = fill_form("fha_max_mortgage_calc", profiles, out, loan_id="12345", run_id="2026-01-01T120000Z")
    assert result["cells_filled"] >= 2  # at least Simple + Streamline
    wb = openpyxl.load_workbook(str(out))
    assert wb["Simple"]["I10"].value == 2100.0
    assert wb[" Streamline (Owner-Occupied)"]["I6"].value == 2100.0


def test_fill_form_writes_numeric_values(tmp_path):
    """Numeric cell types should write Python float/int, not strings."""
    import openpyxl
    profiles = _make_profiles_dir(tmp_path)
    out = tmp_path / "output" / "income_calc_w2.xlsx"
    result = fill_form("income_calc_w2", profiles, out, loan_id="12345", run_id="2026-01-01T120000Z")
    if result["cells_filled"] > 0:
        wb = openpyxl.load_workbook(str(out))
        ws = wb[wb.sheetnames[0]]
        # Check that filled numeric cells have numeric values
        for mapping in FORM_TEMPLATES["income_calc_w2"].mappings:
            sheet = ws if not mapping.sheet else wb[mapping.sheet]
            val = sheet[mapping.cell].value
            if val is not None and mapping.cell_type in (CellType.NUMBER, CellType.CURRENCY, CellType.PERCENTAGE):
                assert isinstance(val, (int, float)), f"Cell {mapping.cell} should be numeric, got {type(val)}"
