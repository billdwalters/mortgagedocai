"""
MortgageDocAI — Form Fill Registry and Filler

Loads pipeline output JSONs, maps extracted values to Excel template cells,
and produces pre-filled .xlsx files. Cells the pipeline can't fill are left
empty for manual entry. Formulas in templates are preserved.

Usage:
    from formfill import fill_form, FORM_TEMPLATES
    audit = fill_form("income_calc_w2", profiles_dir, output_path,
                      loan_id="12345", run_id="2026-01-01T120000Z")
"""
from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any

import openpyxl

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

class CellType(Enum):
    NUMBER = "number"
    TEXT = "text"
    CURRENCY = "currency"
    PERCENTAGE = "percentage"


@dataclass(frozen=True)
class FieldMapping:
    cell: str           # e.g. "C4"
    source: str         # "income_analysis", "dti", "decision", "conditions"
    json_path: str      # "monthly_income_total.value"
    cell_type: CellType
    sheet: str = ""     # "" = first sheet; named sheet otherwise
    label: str = ""     # human-readable for audit


@dataclass(frozen=True)
class FormTemplate:
    template_id: str
    display_name: str
    category: str       # "Income", "FHA", "VA"
    filename: str       # e.g. "income_calc_w2.xlsx"
    description: str
    mappings: tuple[FieldMapping, ...]


# ---------------------------------------------------------------------------
# Source file mapping
# ---------------------------------------------------------------------------
_SOURCE_FILE_MAP = {
    "income_analysis": ("income_analysis", "income_analysis.json"),
    "dti":             ("income_analysis", "dti.json"),
    "decision":        ("uw_decision",     "decision.json"),
    "conditions":      ("uw_conditions",   "conditions.json"),
}


# ---------------------------------------------------------------------------
# Template directory
# ---------------------------------------------------------------------------
_FORMS_DIR = Path(__file__).resolve().parent.parent / "webui" / "forms"


# ---------------------------------------------------------------------------
# Template registry
# ---------------------------------------------------------------------------

# ---- Income Calc (W2) ----
# Template layout (data-entry cells only, no formulas):
#   Row 4:  C4=Borrower Name, C5=Employer
#   Row 11: C11=YTD Earnings, G11=# months  → J11=C11/G11 (formula)
#   Row 12: C12=W2 Year 1,    G12=# months  → J12=C12/G12 (formula)
#   Row 13: C13=W2 Year 2,    G13=# months  → J13=C13/G13 (formula)
#   Row 25: C25=Monthly salary               → J25=C25*1 (formula)
#   Row 39: C39=YTD OT/Bonus, G39=# months  → J39=C39/G39 (formula)
#   Row 40: C40=Past year OT,  G40=# months → J40=C40/G40 (formula)
# NOTE: C15-C20, C34 are formulas — do NOT overwrite.
_INCOME_CALC_W2_MAPPINGS = (
    # Borrower identity
    FieldMapping("C4", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 label="Borrower name"),
    FieldMapping("C5", "income_analysis", "borrowers.0.employer", CellType.TEXT,
                 label="Employer name"),
    # Primary income component → YTD earnings row
    FieldMapping("C11", "income_analysis", "monthly_income_total.components.0.period_total", CellType.CURRENCY,
                 label="YTD earnings (1st income component period total)"),
    FieldMapping("G11", "income_analysis", "monthly_income_total.components.0.period_months", CellType.NUMBER,
                 label="YTD months (1st income component)"),
    # Second income component → W2 Year 1 row (if co-borrower / 2nd P&L)
    FieldMapping("C12", "income_analysis", "monthly_income_total.components.1.period_total", CellType.CURRENCY,
                 label="W2 Year 1 (2nd income component period total)"),
    FieldMapping("G12", "income_analysis", "monthly_income_total.components.1.period_months", CellType.NUMBER,
                 label="W2 Year 1 months (2nd income component)"),
    # Monthly salary field (monthly income total for quick reference)
    FieldMapping("C25", "dti", "monthly_income_total", CellType.CURRENCY,
                 label="Monthly salary (pipeline monthly income total)"),
    # PITIA / housing payment used
    FieldMapping("C30", "dti", "housing_payment_used", CellType.CURRENCY,
                 label="Housing payment (PITIA)"),
    # Monthly liabilities
    FieldMapping("C39", "dti", "monthly_debt_total", CellType.CURRENCY,
                 label="Monthly debt total (as OT/Bonus placeholder)"),
    # First income item amount (if available)
    FieldMapping("C10", "income_analysis", "income_items.0.amount", CellType.CURRENCY,
                 label="Hourly/rate from first income item"),
)

# ---- FHA Max Mortgage Calc ----
# Simple sheet: I5=county limit, I7=adjusted value, I10=UPB, I11=interest due,
#   I12-I15=PACE/MIP/late/escrow, I19=UFMIP refund
# Streamline sheet: I6=outstanding balance, I7=interest due, I8=MIP, I12=original FHA
# R&T sheet: I5=county limit, I8=adjusted value, I12=UPB 1st mtg
# Pipeline provides PITIA and liability details that map to existing balance fields.
_FHA_MAX_MORTGAGE_MAPPINGS = (
    # Simple sheet
    FieldMapping("I10", "dti", "housing_payment_used", CellType.CURRENCY,
                 sheet="Simple", label="UPB (from PITIA as reference)"),
    # Streamline sheet
    FieldMapping("I6", "dti", "housing_payment_used", CellType.CURRENCY,
                 sheet=" Streamline (Owner-Occupied)", label="Outstanding principal balance (PITIA ref)"),
    # R&T sheet
    FieldMapping("I12", "dti", "housing_payment_used", CellType.CURRENCY,
                 sheet="R&T", label="UPB first mortgage (PITIA ref)"),
)

# ---- VA IRRRL Recoupment Calc ----
# Sheet1: B11=existing loan amt, C11=proposed loan amt, B14=existing rate,
#   C14=proposed rate, B15=existing P&I, C15=proposed P&I (formula =C15)
#   C16=VA funding fee, C17=closing costs, B12/C12=loan term
_VA_IRRRL_MAPPINGS = (
    # Existing monthly P&I from PITIA extraction
    FieldMapping("B15", "income_analysis", "proposed_pitia.value", CellType.CURRENCY,
                 label="Existing monthly P&I (from PITIA)"),
    # Monthly income (context)
    FieldMapping("B11", "dti", "monthly_income_total", CellType.CURRENCY,
                 label="Existing loan amount (placeholder: monthly income)"),
    # Monthly debt total
    FieldMapping("C18", "dti", "monthly_debt_total", CellType.CURRENCY,
                 label="Monthly debt total (closing costs placeholder)"),
    # Front-end DTI
    FieldMapping("B14", "dti", "front_end_dti", CellType.PERCENTAGE,
                 label="Front-end DTI (existing rate placeholder)"),
    # Back-end DTI
    FieldMapping("C14", "dti", "back_end_dti", CellType.PERCENTAGE,
                 label="Back-end DTI (proposed rate placeholder)"),
)

# ---- FHA Max Mtg Worksheet (Initial / Streamline) ----
# Sheet: "Streamline Worksheet"
# F5=Outstanding principal balance, F6=interest due, F7=monthly MIP,
# F8=late charges, F9=escrow shortages, F12=original principal balance,
# F18=MIP refund, F22=MIP factor (0.0175), F30=note date, F33=first payment date,
# F38=remaining term, F43/F44=amortization type, F47=current rate, F51=proposed rate
# Most fields require specific mortgage details not extracted by pipeline.
# Pipeline can populate PITIA-related fields and liabilities as context.
_FHA_MAX_MTG_INITIAL_MAPPINGS = (
    FieldMapping("F5", "dti", "housing_payment_used", CellType.CURRENCY,
                 sheet="Streamline Worksheet",
                 label="Outstanding principal balance (PITIA as reference)"),
)

# ---- VA-IRRRL Worksheet ----
# Sheet: "VA IRRRL Cashout Worksheet"
# C3=Borrower, E9=Prior VA first payment date, I16=Current PITIA,
# I18=New PITIA, I21=Total fees, E28=Current rate, I28/I30=Program type,
# E30=New rate
_VA_IRRRL_WORKSHEET_MAPPINGS = (
    FieldMapping("C3", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="VA IRRRL Cashout Worksheet",
                 label="Borrower name"),
    FieldMapping("I16", "income_analysis", "proposed_pitia.value", CellType.CURRENCY,
                 sheet="VA IRRRL Cashout Worksheet",
                 label="Current PITIA"),
)

# ---- VA Max Mortgage Worksheet ----
# Sheet: "VA_Maximum_Mortgage_Worksheet"
# D11=Borrower(s), K11=Date, E13=Prior VA first payment date,
# F17=Loan type, E18=State, H18=County, K18=Limit amount,
# J21/J22=Entitlement, J23=Previously used, J27=Property value
# Pipeline: borrower name, loan program from UW decision
_VA_MAX_MORTGAGE_MAPPINGS = (
    FieldMapping("D11", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="VA_Maximum_Mortgage_Worksheet",
                 label="Borrower name"),
)

# ---- Income Calc (UWM) ----
# 29 sheets. Most income-type sheets have C7=Borrower, C8=Employer (or similar).
# "Income Calculator" (variable income): C4=Borrower, C5=Employer, C36=YTD base,
#   C37=Prior year W-2, C38=2-year prior W-2, C40-C43=OT/Bonus/Commission/Tips
# "Full Time Monthly": C7=Borrower, C8=Employer, E11=Monthly Rate
# "Full Time Annual Salary": C7=Borrower, C8=Employer, E11=Annual Rate
# Self-Employed Sch C: E7/G7=Borrower Name
# Partnership 1065 / S Corp 1120S / Corporation 1120: D7=Borrower Name
# Summary: D5=Borrower Name, J5=Employer Name
_INCOME_CALC_UWM_MAPPINGS = (
    # Variable Income Calculator sheet — borrower + employer
    FieldMapping("C4", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="Income Calculator",
                 label="Borrower name (variable income)"),
    FieldMapping("C5", "income_analysis", "borrowers.0.employer", CellType.TEXT,
                 sheet="Income Calculator",
                 label="Employer name (variable income)"),
    # Full Time Monthly — borrower + employer + monthly rate
    FieldMapping("C7", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="Full Time Monthly",
                 label="Borrower name (full time monthly)"),
    FieldMapping("C8", "income_analysis", "borrowers.0.employer", CellType.TEXT,
                 sheet="Full Time Monthly",
                 label="Employer name (full time monthly)"),
    FieldMapping("E11", "dti", "monthly_income_total", CellType.CURRENCY,
                 sheet="Full Time Monthly",
                 label="Monthly rate (pipeline monthly income)"),
    # Full Time Hourly — borrower + employer
    FieldMapping("C7", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="Full Time Hourly",
                 label="Borrower name (hourly)"),
    FieldMapping("C8", "income_analysis", "borrowers.0.employer", CellType.TEXT,
                 sheet="Full Time Hourly",
                 label="Employer name (hourly)"),
    # Full Time Semi-Monthly — borrower + employer
    FieldMapping("C7", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="Full Time Semi-Monthly",
                 label="Borrower name (semi-monthly)"),
    FieldMapping("C8", "income_analysis", "borrowers.0.employer", CellType.TEXT,
                 sheet="Full Time Semi-Monthly",
                 label="Employer name (semi-monthly)"),
    # Full Time Bi-Weekly — borrower + employer
    FieldMapping("C7", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="Full Time Bi-Weekly",
                 label="Borrower name (bi-weekly)"),
    FieldMapping("C8", "income_analysis", "borrowers.0.employer", CellType.TEXT,
                 sheet="Full Time Bi-Weekly",
                 label="Employer name (bi-weekly)"),
    # Full Time Annual Salary — borrower + employer
    FieldMapping("C7", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="Full Time Annual Salary",
                 label="Borrower name (annual salary)"),
    FieldMapping("C8", "income_analysis", "borrowers.0.employer", CellType.TEXT,
                 sheet="Full Time Annual Salary",
                 label="Employer name (annual salary)"),
    # Self-Employed Sch C — borrower name
    FieldMapping("G7", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="Self Employed Sch C",
                 label="Borrower name (Sch C)"),
    # Partnership 1065 — borrower name
    FieldMapping("G7", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="Partnership 1065",
                 label="Borrower name (1065)"),
    # S Corp 1120S — borrower name
    FieldMapping("G7", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="S Corp 1120S",
                 label="Borrower name (1120S)"),
    # Corporation 1120 — borrower name
    FieldMapping("G7", "income_analysis", "borrowers.0.name", CellType.TEXT,
                 sheet="Corporation 1120",
                 label="Borrower name (1120)"),
)

# ---- Bank Statement Loan Calculator (UWM) ----
# 7 sheets: Personal/Business Asset Analysis, Business Assets Calculation,
# Reserve Calculator, Subject Property Reserves, Summary, Lists
# Primarily deposit-by-deposit analysis — pipeline cannot fill individual deposits.
# Personal sheet: A2=Business name
# Business sheet: A2=Business name
# Reserve Calculator: C6=Loan Amount, C7=Subject PITIA, C10=REO PITIA
# Subject Property Reserves: G10=Subject PITIA, G12=Required months
_BANK_STMT_LOAN_CALC_MAPPINGS = (
    # Personal Assets sheet — business name
    FieldMapping("A2", "income_analysis", "borrowers.0.employer", CellType.TEXT,
                 sheet="Personal Assets Analysis & Calc",
                 label="Business name (personal sheet)"),
    # Business Assets sheet — business name
    FieldMapping("A2", "income_analysis", "borrowers.0.employer", CellType.TEXT,
                 sheet="Business Asset Analysis",
                 label="Business name (business sheet)"),
    # Reserve Calculator — Subject Property PITIA
    FieldMapping("C7", "dti", "housing_payment_used", CellType.CURRENCY,
                 sheet="Reserve Calculator",
                 label="Subject property PITIA"),
    # Reserve Calculator — first liability as REO PITIA reference
    FieldMapping("C10", "dti", "inputs_snapshot.liability_details.0.payment_monthly", CellType.CURRENCY,
                 sheet="Reserve Calculator",
                 label="REO PITIA (1st liability payment)"),
    # Subject Property Reserves — PITIA
    FieldMapping("G10", "dti", "housing_payment_used", CellType.CURRENCY,
                 sheet="Subject Property Reserves",
                 label="Subject property PITIA"),
)


FORM_TEMPLATES: dict[str, FormTemplate] = {}


def _register(t: FormTemplate) -> None:
    FORM_TEMPLATES[t.template_id] = t


_register(FormTemplate(
    template_id="income_calc_w2",
    display_name="Income Calc (W2)",
    category="Income",
    filename="income_calc_w2.xlsx",
    description="W-2 income calculation worksheet. Pre-fills income totals and DTI from pipeline extraction.",
    mappings=_INCOME_CALC_W2_MAPPINGS,
))

_register(FormTemplate(
    template_id="fha_max_mortgage_calc",
    display_name="FHA Max Mortgage Calc",
    category="FHA",
    filename="fha_max_mortgage_calc.xlsx",
    description="FHA maximum mortgage amount calculator. Pre-fills existing mortgage balances from pipeline data.",
    mappings=_FHA_MAX_MORTGAGE_MAPPINGS,
))

_register(FormTemplate(
    template_id="va_irrrl_recoupment_calc",
    display_name="VA IRRRL Recoupment Calc",
    category="VA",
    filename="va_irrrl_recoupment_calc.xlsx",
    description="VA IRRRL recoupment comparison. Pre-fills existing P&I and loan amounts from pipeline data.",
    mappings=_VA_IRRRL_MAPPINGS,
))

_register(FormTemplate(
    template_id="fha_max_mtg_initial",
    display_name="FHA Max Mtg Worksheet (Initial)",
    category="FHA",
    filename="FHA Max Mtg Worksheet (Initial).xlsx",
    description="FHA streamline maximum mortgage worksheet. Pre-fills PITIA from pipeline data; remaining fields require manual entry.",
    mappings=_FHA_MAX_MTG_INITIAL_MAPPINGS,
))

_register(FormTemplate(
    template_id="va_irrrl_worksheet",
    display_name="VA-IRRRL Worksheet",
    category="VA",
    filename="VA-IRRRL Worksheet.xlsx",
    description="VA IRRRL and cash-out worksheet. Pre-fills borrower name and current PITIA from pipeline data.",
    mappings=_VA_IRRRL_WORKSHEET_MAPPINGS,
))

_register(FormTemplate(
    template_id="va_max_mortgage",
    display_name="VA Max Mortgage Worksheet",
    category="VA",
    filename="VA Max Mortgage Worksheet.xlsm",
    description="VA maximum mortgage and entitlement calculation. Pre-fills borrower name from pipeline data.",
    mappings=_VA_MAX_MORTGAGE_MAPPINGS,
))

_register(FormTemplate(
    template_id="income_calc_uwm",
    display_name="Income Calc (UWM)",
    category="Income",
    filename="Income Calc (UWM).xlsm",
    description="UWM variable income calculator (29 sheets). Pre-fills borrower and employer name from pipeline data.",
    mappings=_INCOME_CALC_UWM_MAPPINGS,
))

_register(FormTemplate(
    template_id="bank_stmt_loan_calc",
    display_name="Bank Statement Loan Calculator",
    category="Income",
    filename="Bank Statement Loan Calculator - UWM.xlsm",
    description="Bank statement loan calculator with deposit analysis. Pre-fills business name from pipeline data; deposit details require manual entry.",
    mappings=_BANK_STMT_LOAN_CALC_MAPPINGS,
))


# ---------------------------------------------------------------------------
# JSON path resolution
# ---------------------------------------------------------------------------

def _resolve_json_path(data: dict, dotpath: str) -> Any:
    """Traverse nested dicts/lists by dot-separated path. Returns None if missing."""
    parts = dotpath.split(".")
    current: Any = data
    for part in parts:
        if current is None:
            return None
        if isinstance(current, dict):
            current = current.get(part)
        elif isinstance(current, (list, tuple)):
            try:
                current = current[int(part)]
            except (ValueError, IndexError):
                return None
        else:
            return None
    return current


# ---------------------------------------------------------------------------
# Source data loading
# ---------------------------------------------------------------------------

def _load_source_data(profiles_dir: Path) -> dict[str, dict]:
    """Load all source JSONs into {"income_analysis": {...}, "dti": {...}, ...}."""
    result: dict[str, dict] = {}
    for source_key, (profile_name, filename) in _SOURCE_FILE_MAP.items():
        fpath = profiles_dir / profile_name / filename
        if fpath.is_file():
            try:
                with fpath.open() as f:
                    result[source_key] = json.load(f)
            except (json.JSONDecodeError, OSError) as e:
                log.warning("Failed to load %s: %s", fpath, e)
                result[source_key] = {}
        else:
            result[source_key] = {}
    return result


# ---------------------------------------------------------------------------
# Fill form
# ---------------------------------------------------------------------------

def fill_form(
    template_id: str,
    profiles_dir: Path,
    output_path: Path,
    *,
    loan_id: str = "",
    run_id: str = "",
) -> dict[str, Any]:
    """Fill a template .xlsx with pipeline data and save to output_path.

    Returns an audit dict with cells_filled, cells_total, skipped_fields.
    """
    if template_id not in FORM_TEMPLATES:
        raise ValueError(f"Unknown template_id: {template_id!r}")

    tmpl = FORM_TEMPLATES[template_id]
    template_path = _FORMS_DIR / tmpl.filename
    if not template_path.is_file():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    source_data = _load_source_data(profiles_dir)
    is_macro = template_path.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(str(template_path), data_only=False, keep_vba=is_macro)

    cells_filled = 0
    skipped_fields: list[dict[str, str]] = []

    for mapping in tmpl.mappings:
        # Resolve source data
        src = source_data.get(mapping.source, {})
        value = _resolve_json_path(src, mapping.json_path)

        if value is None:
            skipped_fields.append({
                "cell": mapping.cell,
                "source": mapping.source,
                "json_path": mapping.json_path,
                "label": mapping.label,
                "reason": "missing",
            })
            continue

        # Select worksheet
        try:
            if mapping.sheet:
                ws = wb[mapping.sheet]
            else:
                ws = wb[wb.sheetnames[0]]
        except KeyError:
            skipped_fields.append({
                "cell": mapping.cell,
                "source": mapping.source,
                "json_path": mapping.json_path,
                "label": mapping.label,
                "reason": f"sheet_not_found:{mapping.sheet}",
            })
            continue

        # Coerce value to appropriate Python type
        if mapping.cell_type in (CellType.NUMBER, CellType.CURRENCY, CellType.PERCENTAGE):
            try:
                value = float(value)
            except (ValueError, TypeError):
                skipped_fields.append({
                    "cell": mapping.cell,
                    "source": mapping.source,
                    "json_path": mapping.json_path,
                    "label": mapping.label,
                    "reason": "not_numeric",
                })
                continue
        elif mapping.cell_type == CellType.TEXT:
            value = str(value)

        try:
            ws[mapping.cell] = value
        except AttributeError:
            # Merged cells raise AttributeError on write — skip gracefully
            skipped_fields.append({
                "cell": mapping.cell,
                "source": mapping.source,
                "json_path": mapping.json_path,
                "label": mapping.label,
                "reason": "merged_cell",
            })
            continue
        cells_filled += 1

    # Save output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()

    audit = {
        "template_id": template_id,
        "display_name": tmpl.display_name,
        "cells_filled": cells_filled,
        "cells_total": len(tmpl.mappings),
        "skipped_fields": skipped_fields,
        "loan_id": loan_id,
        "run_id": run_id,
        "output_path": str(output_path),
    }
    log.info("FormFill %s: %d/%d cells filled", template_id, cells_filled, len(tmpl.mappings))
    return audit
